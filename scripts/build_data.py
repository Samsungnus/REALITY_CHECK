from __future__ import annotations

import json
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "source-data"
OUTPUT_DIR = ROOT / "src" / "data"

PRICE_FILE = SOURCE_DIR / "bratislava_ceny_bytov_2020_2026.xlsx"
APARTMENT_FILE = SOURCE_DIR / "developerske-projekty.xlsx"
DEVELOPER_FILE = SOURCE_DIR / "Developers.xlsx"
DEVELOPER_SOURCE_URL = "https://github.com/Samsungnus/REALITY_CHECK/blob/main/source-data/Developers.xlsx"

DEVELOPER_FALLBACKS = {
    "Slnečná Strana": "Slnečná Strana",
}


def serialise(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def find_header_row(sheet, required: set[str], scan_limit: int = 30) -> int:
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True),
        start=1,
    ):
        values = {str(value).strip() for value in row if value is not None}
        if required.issubset(values):
            return row_number
    raise ValueError(
        f"V liste {sheet.title!r} sa nenašla hlavička s poľami: {', '.join(sorted(required))}"
    )


def extract_table(sheet, required: set[str]) -> tuple[list[str], list[dict[str, Any]]]:
    header_row = find_header_row(sheet, required)
    raw_headers = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    )
    headers = [str(value).strip() if value is not None else f"Stĺpec {index}" for index, value in enumerate(raw_headers, start=1)]

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None and value != "" for value in values):
            continue
        rows.append(
            {
                header: serialise(value)
                for header, value in zip(headers, values)
            }
        )
    return headers, rows


def write_json(filename: str, payload: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with (OUTPUT_DIR / filename).open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, allow_nan=False)


def main() -> None:
    for path in (PRICE_FILE, APARTMENT_FILE, DEVELOPER_FILE):
        if not path.exists():
            raise FileNotFoundError(f"Chýba zdrojový súbor: {path}")

    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    price_workbook = openpyxl.load_workbook(PRICE_FILE, data_only=True, read_only=True)
    if "Data" not in price_workbook.sheetnames:
        raise ValueError("Cenový Excel musí obsahovať list 'Data'.")
    price_headers, price_rows = extract_table(
        price_workbook["Data"], {"Dátum", "Mesiac", "Novostavby €/m²"}
    )
    write_json(
        "prices.json",
        {
            "meta": {
                "sourceFile": PRICE_FILE.name,
                "sheet": "Data",
                "generatedAt": generated_at,
                "rowCount": len(price_rows),
                "columnCount": len(price_headers),
            },
            "columns": price_headers,
            "rows": price_rows,
        },
    )

    source_rows: list[dict[str, Any]] = []
    if "Zdroje" in price_workbook.sheetnames:
        _, source_rows = extract_table(
            price_workbook["Zdroje"], {"Segment", "Obdobie", "URL"}
        )
    source_rows.append(
        {
            "Segment": "Developeri",
            "Obdobie": "aktuálne",
            "Hodnota / informácia": "Podmienky kúpy u sledovaných developerov",
            "Použitie": "Parkovanie, kobky, spoločné priestory, financovanie a kolaudácia",
            "URL": DEVELOPER_SOURCE_URL,
        }
    )
    write_json(
        "sources.json",
        {
            "meta": {
                "sourceFile": PRICE_FILE.name,
                "sheet": "Zdroje",
                "generatedAt": generated_at,
                "rowCount": len(source_rows),
            },
            "rows": source_rows,
        },
    )
    price_workbook.close()

    apartment_workbook = openpyxl.load_workbook(
        APARTMENT_FILE, data_only=True, read_only=True
    )
    if "Byty" not in apartment_workbook.sheetnames:
        raise ValueError("Developerský Excel musí obsahovať list 'Byty'.")
    apartment_headers, apartment_rows = extract_table(
        apartment_workbook["Byty"], {"Projekt", "Označenie bytu", "Stav"}
    )
    for row in apartment_rows:
        if not str(row.get("Developer") or "").strip():
            fallback = DEVELOPER_FALLBACKS.get(str(row.get("Projekt") or "").strip())
            if fallback:
                row["Developer"] = fallback
    write_json(
        "apartments.json",
        {
            "meta": {
                "sourceFile": APARTMENT_FILE.name,
                "sheet": "Byty",
                "generatedAt": generated_at,
                "rowCount": len(apartment_rows),
                "columnCount": len(apartment_headers),
            },
            "columns": apartment_headers,
            "rows": apartment_rows,
        },
    )
    apartment_workbook.close()

    developer_workbook = openpyxl.load_workbook(
        DEVELOPER_FILE, data_only=True, read_only=True
    )
    if "Projekty" not in developer_workbook.sheetnames:
        raise ValueError("Excel developerov musí obsahovať list 'Projekty'.")
    _, developer_project_rows = extract_table(
        developer_workbook["Projekty"],
        {"Developer", "Názov projektu", "Parkovanie", "Zdroj / overenie"},
    )

    tracked_projects = {
        str(row.get("Projekt") or "").strip(): str(row.get("Developer") or "").strip()
        for row in apartment_rows
        if str(row.get("Projekt") or "").strip()
    }
    detail_fields = {
        "parkingRequired": "Parkovanie",
        "parkingPrice": "Cena parkovania",
        "storageRequired": "Kobka",
        "storagePrice": "Cena kobky",
        "commonSpace": "Spoločný priestor",
        "bikeSpace": "Priestor na bicykle",
        "financing": "Financovanie",
        "largePaymentWhen": "Vyššia časť sa platí",
        "completion": "Kolaudácia",
    }
    project_details: dict[str, dict[str, Any]] = {}
    grouped_details: dict[str, list[tuple[str, dict[str, Any]]]] = {}

    for source_row in developer_project_rows:
        project = str(source_row.get("Názov projektu") or "").strip()
        display_developer = tracked_projects.get(project)
        if not project or not display_developer:
            continue
        details = {
            target: source_row.get(source)
            for target, source in detail_fields.items()
        }
        raw_sources = str(source_row.get("Zdroj / overenie") or "")
        details["sourceUrls"] = [
            url.strip() for url in raw_sources.split("|") if url.strip()
        ]
        project_details[project] = details
        grouped_details.setdefault(display_developer, []).append((project, details))

    developer_details: dict[str, dict[str, Any]] = {}
    for developer, entries in grouped_details.items():
        combined: dict[str, Any] = {}
        for field in detail_fields:
            if len(entries) == 1:
                combined[field] = entries[0][1].get(field)
            else:
                combined[field] = " | ".join(
                    f"{project}: {details.get(field) or 'Nezistené'}"
                    for project, details in entries
                )
        combined["sourceUrls"] = list(
            dict.fromkeys(
                url
                for _, details in entries
                for url in details.get("sourceUrls", [])
            )
        )
        combined["note"] = (
            "Podmienky sú uvedené osobitne podľa projektu."
            if len(entries) > 1
            else None
        )
        developer_details[developer] = combined

    write_json(
        "developers.json",
        {
            "meta": {
                "sourceFile": DEVELOPER_FILE.name,
                "sheet": "Projekty",
                "generatedAt": generated_at,
                "rowCount": len(project_details),
            },
            "developers": developer_details,
            "projects": project_details,
        },
    )
    developer_workbook.close()

    print(
        f"Pripravené: {len(price_rows)} cenových záznamov, "
        f"{len(apartment_rows)} bytov, {len(source_rows)} zdrojov a "
        f"{len(project_details)} projektových podmienok."
    )


if __name__ == "__main__":
    main()

